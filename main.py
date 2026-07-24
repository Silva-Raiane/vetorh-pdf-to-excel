"""
Senior VetorRH - Extrator de Apuração de Ponto (PDF para Excel)

Este módulo extrai registros de ocorrência e apuração de ponto de relatórios PDF 
gerados pelo sistema Senior VetorRH e consolida os dados em uma planilha Excel.
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# Configurações e Constantes
# ----------------------------------------------------------------------
CODIGOS_SITUACAO = {
    "024", "100", "115", "120", "135", 
    "200", "250", "300", "301", "999"
}

NOISE_PATTERNS = [
    re.compile(r"^Pag\.:\s*\d+$"),
    re.compile(r"^Apuração Colaborador$"),
    re.compile(r"^Período de:"),
    re.compile(r"^HRAP110\.APU"),
]

HEADER_RE = re.compile(r"^(\d{5})\s+(.+?)\s+(\d{4})$")
DATE_RE = re.compile(r"^(\d{2}/\d{2}/\d{2})\s+(\w{3})\s+(.*)$")
TOTAL_HORAS_RE = re.compile(r"^\d{3,}:\d{2}$")


# ----------------------------------------------------------------------
# 1. Extração do texto do PDF
# ----------------------------------------------------------------------
def extrair_texto(caminho_pdf: Path) -> str:
    """Extrai e concatena o texto de todas as páginas do PDF."""
    if not caminho_pdf.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho_pdf}")

    with pdfplumber.open(caminho_pdf) as pdf:
        paginas = [pagina.extract_text() or "" for pagina in pdf.pages]
    return "\n".join(paginas)


# ----------------------------------------------------------------------
# 2. Parsing do texto para uma lista de registros
# ----------------------------------------------------------------------
def _is_noise(linha: str) -> bool:
    """Verifica se a linha corresponde a um ruído/cabeçalho de página."""
    return any(p.match(linha) for p in NOISE_PATTERNS)


def parse_relatorio(texto: str) -> List[Dict[str, object]]:
    """Realiza o parse do texto bruto extraído do PDF e retorna uma lista de dicionários."""
    linhas = [l.strip() for l in texto.split("\n") if l.strip()]

    registros = []
    matricula_atual: Optional[str] = None
    nome_atual: Optional[str] = None
    em_totais = False
    ultimo_dia = ""
    ultima_marcacao = ""

    for linha in linhas:
        if _is_noise(linha):
            continue

        m = HEADER_RE.match(linha)
        if m:
            matricula, nome, _codigo_turno = m.groups()
            if matricula != matricula_atual:
                matricula_atual, nome_atual = matricula, nome
                ultimo_dia = ultima_marcacao = ""
            em_totais = False
            continue

        if linha == "Dia Marcações Situações Apuradas Horas":
            continue

        if linha.startswith("Total Colaborador:"):
            em_totais = True
            continue

        tokens = linha.split()
        if tokens and TOTAL_HORAS_RE.match(tokens[-1]):
            continue
        if em_totais or not matricula_atual:
            continue

        dm = DATE_RE.match(linha)
        tem_data_propria = bool(dm)
        if dm:
            data_bruta, _dow, resto = dm.groups()
            d, mo, y = data_bruta.split("/")
            dia = f"{d}/{mo}/20{y}"
            ultimo_dia = dia
        else:
            resto = linha
            dia = ultimo_dia

        tokens = resto.split()
        idx_codigo = next((i for i, t in enumerate(tokens) if t in CODIGOS_SITUACAO), None)
        if idx_codigo is None:
            continue

        marc_tokens = tokens[:idx_codigo]
        codigo = tokens[idx_codigo]
        desc_tokens = tokens[idx_codigo + 1 : -1]
        horas = tokens[-1]

        if marc_tokens:
            marcacoes = " ".join(marc_tokens)
            ultima_marcacao = marcacoes
        elif tem_data_propria:
            marcacoes = ""
            ultima_marcacao = ""
        else:
            marcacoes = ultima_marcacao

        registros.append(
            {
                "MATRICULA": int(matricula_atual),
                "EMPREGADO": nome_atual,
                "DIA": dia,
                "MARCAÇÕES": marcacoes,
                "SITUAÇÕES APURADAS": f"{codigo} {' '.join(desc_tokens)}",
                "HORAS": horas,
            }
        )

    return registros


# ----------------------------------------------------------------------
# 3. Geração do Excel
# ----------------------------------------------------------------------
def gerar_excel(registros: List[Dict[str, object]], caminho_saida: Path) -> None:
    """Gera um arquivo Excel formatado a partir dos registros extraídos."""
    colunas = ["MATRICULA", "EMPREGADO", "DIA", "MARCAÇÕES", "SITUAÇÕES APURADAS", "HORAS"]
    df = pd.DataFrame(registros, columns=colunas)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Apuracao")
        ws = writer.sheets["Apuracao"]

        # Estilização do Cabeçalho
        fonte_cabecalho = Font(name="Arial", bold=True, color="FFFFFF")
        preenchimento = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cel in ws[1]:
            cel.font = fonte_cabecalho
            cel.fill = preenchimento
            cel.alignment = Alignment(horizontal="center")

        # Fonte das células de dados
        for linha in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cel in linha:
                cel.font = Font(name="Arial")

        # Larguras personalizadas das colunas
        larguras = {"A": 12, "B": 38, "C": 13, "D": 22, "E": 34, "F": 10}
        for col, largura in larguras.items():
            ws.column_dimensions[col].width = largura

        # Recursos adicionais de usabilidade
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


# ----------------------------------------------------------------------
# 4. Ponto de Entrada (CLI)
# ----------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Extrai apuração de ponto do Senior VetorRH em PDF para Excel."
    )
    parser.add_argument(
        "-i", "--input",
        type=Path,
        default=Path("HRAP110_TMPFD40.PDF"),
        help="Caminho do arquivo PDF de entrada (Padrão: HRAP110_TMPFD40.PDF)"
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=Path("saida.xlsx"),
        help="Caminho do arquivo Excel de saída (Padrão: saida.xlsx)"
    )

    args = parser.parse_args()

    try:
        print(f"Lendo arquivo PDF: {args.input}...")
        texto = extrair_texto(args.input)

        print("Processando e estruturando os dados...")
        registros = parse_relatorio(texto)

        if not registros:
            print("Nenhum registro correspondente aos códigos cadastrados foi encontrado.")
            sys.exit(0)

        print(f"Gerando planilha Excel: {args.output}...")
        gerar_excel(registros, args.output)

        total_colaboradores = len({r["MATRICULA"] for r in registros})
        print(f"Concluído! {len(registros)} linhas geradas para {total_colaboradores} colaboradores.")

    except Exception as e:
        print(f"Erro durante a execução: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()